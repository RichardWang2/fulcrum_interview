from openpyxl import load_workbook
from openai import OpenAI
import pandas as pd
from typing import List, Tuple, Dict
import os
from dotenv import load_dotenv
from pathlib import Path
import json
# Load the .env file
load_dotenv()

class LLMExcelAnalyzer:
    def __init__(self, api_key: str = None):
        self.api_key = api_key or os.getenv('OPENAI_API_KEY')
        if not self.api_key:
            raise ValueError("OpenAI API key is required. Set it in .env file or pass it to the constructor.")
        self.client = OpenAI(api_key=self.api_key)

    def _is_row_empty(self, row: Tuple) -> bool:
        """Check if a row is empty or contains only None/empty values"""
        return all(cell is None or str(cell).strip() == '' for cell in row)

    def _find_tables(self, sheet) -> List[Dict]:
        """
        Identify tables in the sheet by looking for clusters of non-empty rows
        Returns list of table information including boundaries and data
        """
        tables = []
        current_table_start = None
        all_rows = list(sheet.iter_rows(values_only=True))
        
        for idx, row in enumerate(all_rows):
            is_empty = self._is_row_empty(row)
            
            if not is_empty and current_table_start is None:
                current_table_start = idx
            elif is_empty and current_table_start is not None:
                table_data = all_rows[current_table_start+1:idx]  # Skip the first row
                if len(table_data) >= 2:
                    tables.append({
                        'start_row': current_table_start+1,  # Adjusted start row
                        'end_row': idx,
                        'data': pd.DataFrame(table_data)
                    })
                current_table_start = None
        
        if current_table_start is not None:
            table_data = all_rows[current_table_start+1:]  # Skip the first row
            if len(table_data) >= 2:
                tables.append({
                    'start_row': current_table_start+1,  # Adjusted start row
                    'end_row': len(all_rows),
                    'data': pd.DataFrame(table_data)
                })
        
        return tables

    def _identify_similar_columns(self, tables: List[Dict]) -> Dict[str, str]:
        """
        Use GPT-4 to identify columns that semantically mean the same thing
        Returns a mapping of original column names to standardized names
        """
        # Collect all unique column names from all tables
        all_columns = set()
        for table in tables:
            print(table['data'])
            df = table['data']
            # Convert all column names to strings
            columns = [str(col) for col in df.columns]
            all_columns.update(columns)
        
        # Debug: Print all unique columns found
        print("\nDEBUG - All unique columns found:")
        for col in sorted(all_columns):
            print(f"  - {col}")
        
        prompt = f"""
        Here are column names from multiple benefits data tables:
        {sorted(list(all_columns))}

        Identify columns that semantically mean the same thing but use different names.
        Group them together and suggest a standardized name for each group.

        Return the result as a JSON dictionary mapping original column names to standardized names.

        Rules:
        1. Use lowercase with underscores for standardized names
        2. Only include columns that are part of a similar group
        3. Return a valid JSON dictionary

        Example:
        If you see these columns:
        - "DOB", "Birth Date", "Date of Birth"
        - "EE Only", "Single", "Employee"

        Return exactly this format:
        {{
            "DOB": "date_of_birth",
            "Birth Date": "date_of_birth",
            "Date of Birth": "date_of_birth",
            "EE Only": "employee_only",
            "Single": "employee_only",
            "Employee": "employee_only"
        }}
        """

        try:
            # Debug: Print the prompt being sent to GPT
            print("\nDEBUG - Prompt being sent to GPT:")
            print("-" * 80)
            print(prompt)
            print("-" * 80)

            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a data analyst who identifies semantically similar columns. Always respond with a valid JSON dictionary."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                response_format={"type": "json_object"}
            )
            
            # Debug: Print the raw response
            print("\nDEBUG - Raw GPT response:")
            print("-" * 80)
            print(response.choices[0].message.content)
            print("-" * 80)
            
            # Parse the response as a dictionary
            return json.loads(response.choices[0].message.content)

        except Exception as e:
            print(f"Error identifying similar columns: {str(e)}")
            print("Raw response:", response.choices[0].message.content if 'response' in locals() else "No response")
            return {}

    def analyze_directory(self, directory_path: str, file_pattern: str = "*.xlsx") -> List[Dict]:
        """
        Analyze all Excel files in a directory
        """
        results = []
        directory = Path(directory_path)
        excel_files = list(directory.glob(file_pattern))
        
        print(f"\nFound {len(excel_files)} Excel files in {directory_path}")
        
        # First pass: collect all tables
        all_tables = []
        for file_path in excel_files:
            try:
                print(f"\nReading: {file_path.name}")
                wb = load_workbook(str(file_path))
                sheet = wb.active
                tables = self._find_tables(sheet)
                all_tables.extend(tables)
                
            except Exception as e:
                print(f"Error reading {file_path.name}: {str(e)}")
        
        # Identify similar columns
        print("\nAnalyzing column similarities...")
        column_mapping = self._identify_similar_columns(all_tables)
        
        if column_mapping:
            print("\nFound similar columns:")
            # Group by standardized name for clearer output
            by_standard = {}
            for orig, std in column_mapping.items():
                if std not in by_standard:
                    by_standard[std] = []
                by_standard[std].append(orig)
            
            for std_name, orig_names in by_standard.items():
                print(f"\n{std_name}:")
                for orig in orig_names:
                    print(f"  - {orig}")
            
            # Apply the mapping to all tables
            for table in all_tables:
                table['data'] = table['data'].rename(columns=column_mapping)
        else:
            print("\nNo similar columns found or error in analysis.")
        
        # Print results
        print(f"\nProcessed {len(all_tables)} tables total")
        
        return all_tables

def main():
    try:
        directory_path = "files"  # Replace with your directory path
        analyzer = LLMExcelAnalyzer()
        analyzer.analyze_directory(directory_path)
            
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main() 